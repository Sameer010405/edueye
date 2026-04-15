import os
import sys
import cv2
import dlib
import numpy as np
import time
import pickle
import queue
import threading
import gc
import traceback
from datetime import datetime, date

# TensorFlow Keras compatibility mode (required for DeepFace)
os.environ['TF_USE_LEGACY_KERAS'] = '1'

# DeepFace and Mediapipe imports
DEEPFACE_AVAILABLE = False
EMOTION_LABELS = ['angry', 'disgust', 'fear', 'happy', 'sad', 'surprise', 'neutral']
try:
    from deepface import DeepFace
    DEEPFACE_AVAILABLE = True
except ImportError:
    print("WARNING: DeepFace library not found. Install with: pip install deepface")
except Exception as e:
    print(f"WARNING: Error importing DeepFace: {e}")
    traceback.print_exc(limit=1)

BLAZEFACE_AVAILABLE = False
try:
    import mediapipe as mp
    BLAZEFACE_AVAILABLE = True
except ImportError:
    print("WARNING: Mediapipe library not found. Install with: pip install mediapipe")
except Exception as e:
    print(f"WARNING: Error importing Mediapipe: {e}")
    traceback.print_exc(limit=1)

import face_recognition

# For saving Excel reports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("WARNING: openpyxl library not found. Install with: pip install openpyxl")
    OPENPYXL_AVAILABLE = False

# --- PyQt6 Imports ---
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QVBoxLayout,
    QHBoxLayout, QGridLayout, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QSizePolicy
)
from PyQt6.QtGui import QPixmap, QImage, QFont, QColor, QPainter, QBrush, QPen
from PyQt6.QtCore import Qt, QTimer, QSize, QRect, QPoint, QThread, pyqtSignal

# --- Configuration ---
# Paths are resolved relative to this script's location so the project
# works on any machine without manual path editing.
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

KNOWN_FACES_DIR      = os.path.join(_BASE_DIR, "known_faces")
BASE_LOG_DIR         = _BASE_DIR
ENCODINGS_FILE_PATH  = os.path.join(_BASE_DIR, "known_face_encodings.pkl")
ATTENDANCE_DAY_DIR   = os.path.join(_BASE_DIR, "attendance_day")

LAPTOP_CAM_SRC = 0
IPHONE_CAM_SRC_URL = "http://192.168.1.100:4747/video"  # Change to your phone's IP
IPHONE_CAM_SRC = IPHONE_CAM_SRC_URL

REQUEST_RESOLUTION = (1280, 720)
REQUESTED_FACE_DETECT_MODEL = "hog"
RECOGNITION_THRESHOLD = 0.55
FACE_DETECT_FRAME_SKIP = 10
UNKNOWN_FACE_LABEL = "Unknown"
TRACKER_CONFIDENCE_THRESHOLD = 8.5
MAX_TRACKER_STALE_CYCLES = 3

# DeepFace configurations
DEEPFACE_DETECTOR_BACKEND_GENERAL_TEST = 'opencv'
EMOTION_DETECTOR_FOR_DEEPFACE = 'opencv'
DEEPFACE_ENFORCE_DETECTION = False
EMOTION_ANALYZE_FRAME_SKIP = 10
DEFAULT_EMOTION_STATE  = "Analyzing..."
EMOTION_ERROR_STATE    = "N/A"
EMOTION_DISABLED_STATE = "Disabled"
EMOTION_OFF_STATE      = "Off"

NUM_LECTURES = 8
THUMBNAIL_SIZE_QT = QSize(45, 45)
WINDOW_TITLE_QT = "EDUEYE — Attendance & Emotion System"

DETAIL_WINDOW_WIDTH_QT   = 350
DETAIL_WINDOW_HEIGHT_QT  = 280
DETAIL_BG_COLOR_QT       = QColor(30, 30, 40)
DETAIL_HEADER_COLOR_QT   = QColor(220, 220, 255)
DETAIL_DOMINANT_COLOR_QT = QColor(0, 200, 100)
DETAIL_PERCENT_COLOR_QT  = QColor(180, 180, 200)
DETAIL_BAR_COLOR_QT      = QColor(100, 100, 130)
DETAIL_BAR_OUTLINE_COLOR_QT = QColor(60, 60, 80)
DETAIL_FONT_FAMILY = "Segoe UI"

# --- Global Queues ---
identification_queue = queue.Queue(maxsize=100)
results_queue = queue.Queue()

# ==============================================================================
# Camera Worker Thread
# ==============================================================================
class CameraWorker(QThread):
    camera_opened_signal = pyqtSignal(bool, object, int, int)
    camera_error_signal  = pyqtSignal(str)

    def __init__(self, source, request_resolution):
        super().__init__()
        self.source = source
        self.request_resolution = request_resolution
        self.cap = None
        self._is_running = True

    def run(self):
        print(f"  CameraWorker: Attempting to open source: {self.source}")
        cap_obj = None
        try:
            api_pref = cv2.CAP_ANY
            if isinstance(self.source, int) and os.name == 'nt':
                api_pref = cv2.CAP_DSHOW
            elif isinstance(self.source, str) and self.source.startswith(("http", "rtsp")):
                api_pref = cv2.CAP_FFMPEG

            cap_obj = cv2.VideoCapture(self.source, api_pref)

            if not self._is_running:
                if cap_obj and cap_obj.isOpened():
                    cap_obj.release()
                return

            if not cap_obj.isOpened():
                if api_pref != cv2.CAP_ANY:
                    cap_obj = cv2.VideoCapture(self.source, cv2.CAP_ANY)
                    if not self._is_running:
                        if cap_obj and cap_obj.isOpened():
                            cap_obj.release()
                        return
                    if not cap_obj.isOpened():
                        raise IOError(f"Cannot open video source: {self.source}")
                else:
                    raise IOError(f"Cannot open video source: {self.source}")

            w_req, h_req = self.request_resolution
            cap_obj.set(cv2.CAP_PROP_FRAME_WIDTH,  float(w_req))
            cap_obj.set(cv2.CAP_PROP_FRAME_HEIGHT, float(h_req))
            time.sleep(0.5)

            if not self._is_running:
                if cap_obj and cap_obj.isOpened():
                    cap_obj.release()
                return

            frame_width  = int(cap_obj.get(cv2.CAP_PROP_FRAME_WIDTH))
            frame_height = int(cap_obj.get(cv2.CAP_PROP_FRAME_HEIGHT))

            if frame_width == 0 or frame_height == 0:
                ret_test, _ = cap_obj.read()
                if not self._is_running:
                    if cap_obj and cap_obj.isOpened():
                        cap_obj.release()
                    return
                if not ret_test:
                    raise IOError(f"Failed to read test frame from {self.source}.")
                frame_width  = int(cap_obj.get(cv2.CAP_PROP_FRAME_WIDTH))
                frame_height = int(cap_obj.get(cv2.CAP_PROP_FRAME_HEIGHT))
                if frame_width == 0 or frame_height == 0:
                    raise IOError(f"Still 0 resolution from {self.source} after test read.")

            self.cap = cap_obj
            self.camera_opened_signal.emit(True, self.cap, frame_width, frame_height)
            print(f"  CameraWorker: Opened {self.source} at {frame_width}x{frame_height}")

        except Exception as e:
            msg = f"Failed to open camera '{self.source}': {type(e).__name__} - {e}"
            print(f"  CameraWorker ERROR: {msg}")
            traceback.print_exc(limit=1)
            if cap_obj and cap_obj.isOpened():
                cap_obj.release()
            self.cap = None
            if self._is_running:
                self.camera_error_signal.emit(msg)
        finally:
            if not self._is_running and cap_obj and (self.cap is None or self.cap != cap_obj):
                if cap_obj.isOpened():
                    cap_obj.release()

    def stop(self):
        self._is_running = False
        if self.cap:
            try:
                self.cap.release()
            except Exception as e:
                print(f"  CameraWorker: Exception releasing cap: {e}")
            self.cap = None
        if self.isRunning():
            self.quit()
            if not self.wait(2000):
                self.terminate()
                self.wait()


# ==============================================================================
# IntegratedAttendanceEmotionSystem
# ==============================================================================
class IntegratedAttendanceEmotionSystem:
    def __init__(self, known_faces_dir, recognition_threshold=RECOGNITION_THRESHOLD):
        self.known_faces_dir = known_faces_dir
        self.recognition_threshold = recognition_threshold

        self.known_face_encodings = []
        self.known_face_names = []
        self.person_thumbnails_paths = {}
        self.all_known_person_names = set()

        self.active_trackers  = {}
        self.tracker_id_counter = 0
        self.tracker_details  = {}

        self.frame_count = 0
        self.current_lecture = 1
        self.daily_attendance = {}
        self.today_date = date.today()

        self.display_emotion_details_globally = False
        self.emotion_detection_active = False
        self.deepface_functional = False

        print("=" * 30 + " Initializing System " + "=" * 30)

        self.face_detect_model_to_use = REQUESTED_FACE_DETECT_MODEL
        self.blazeface_detector = None
        self.mp_face_detection_solution = None

        if REQUESTED_FACE_DETECT_MODEL == "blazeface":
            if BLAZEFACE_AVAILABLE and mp:
                try:
                    self.mp_face_detection_solution = mp.solutions.face_detection
                    self.blazeface_detector = self.mp_face_detection_solution.FaceDetection(
                        model_selection=0,
                        min_detection_confidence=0.55
                    )
                    print("INFO: BlazeFace detector initialized.")
                except Exception as e:
                    print(f"ERROR: BlazeFace init failed: {e}. Falling back to 'hog'.")
                    traceback.print_exc(limit=1)
                    self.face_detect_model_to_use = "hog"
            else:
                print("WARNING: BlazeFace selected but Mediapipe not available. Falling back to 'hog'.")
                self.face_detect_model_to_use = "hog"

        print(f"Face detect model in use: {self.face_detect_model_to_use.upper()}")

        self._validate_paths()

        if self.face_detect_model_to_use == "cnn":
            self._check_dlib_cuda()

        if DEEPFACE_AVAILABLE:
            self.deepface_functional = self._initialize_deepface()
        else:
            print("INFO: DeepFace not available. Emotion detection disabled.")

        self.load_known_faces()
        self._initialize_attendance()
        self.start_identification_thread()
        print("=" * 80 + "\nSYSTEM READY\n" + "=" * 80)

    def _check_dlib_cuda(self):
        print("--- Checking dlib CUDA Status ---")
        try:
            if dlib.DLIB_USE_CUDA:
                num_devices = dlib.cuda.get_num_devices()
                print(f"INFO: dlib CUDA enabled. Devices found: {num_devices}")
            else:
                print("WARNING: dlib not compiled with CUDA (CNN model will be slow).")
        except AttributeError:
            print("WARNING: Cannot determine dlib CUDA status.")
        except Exception as e:
            print(f"WARNING: Error checking dlib CUDA: {e}")

    def _validate_paths(self):
        print("--- Validating Paths ---")
        try:
            os.makedirs(ATTENDANCE_DAY_DIR, exist_ok=True)
            print(f"OK: attendance_day directory: '{ATTENDANCE_DAY_DIR}'")
        except OSError as e:
            errmsg = f"FATAL ERROR: Cannot create attendance directory:\n{ATTENDANCE_DAY_DIR}\n{e}"
            print(errmsg)
            if QApplication.instance():
                QMessageBox.critical(None, "Configuration Error", errmsg)
            sys.exit(1)

    def _initialize_deepface(self):
        if not DEEPFACE_AVAILABLE:
            return False
        print("--- Initializing DeepFace ---")
        start_time = time.time()
        dummy_frame = np.zeros((224, 224, 3), dtype=np.uint8)
        cv2.rectangle(dummy_frame, (50, 50), (174, 174), (200, 200, 200), -1)

        def _test_backend(backend):
            try:
                DeepFace.analyze(
                    img_path=dummy_frame.copy(),
                    actions=['emotion'],
                    detector_backend=backend,
                    enforce_detection=DEEPFACE_ENFORCE_DETECTION,
                    silent=True
                )
                return True
            except ValueError as e:
                if "Face could not be detected" in str(e) and not DEEPFACE_ENFORCE_DETECTION:
                    return True  # Models loaded; no face in dummy is expected
                return False
            except Exception:
                return False

        ok_general = _test_backend(DEEPFACE_DETECTOR_BACKEND_GENERAL_TEST)
        ok_emotion  = (
            _test_backend(EMOTION_DETECTOR_FOR_DEEPFACE)
            if EMOTION_DETECTOR_FOR_DEEPFACE != DEEPFACE_DETECTOR_BACKEND_GENERAL_TEST
            else ok_general
        )
        elapsed = time.time() - start_time
        status = ok_general and ok_emotion
        print(f"DeepFace init {'OK' if status else 'FAILED'} ({elapsed:.2f}s)")
        return status

    def load_known_faces(self):
        print(f"--- Loading face encodings from '{ENCODINGS_FILE_PATH}' ---")
        if not os.path.exists(ENCODINGS_FILE_PATH):
            errmsg = (
                f"FATAL: Encodings file not found:\n{ENCODINGS_FILE_PATH}\n"
                "Run encode_face_v2.py first."
            )
            print(errmsg)
            if QApplication.instance():
                QMessageBox.critical(None, "Initialization Error", errmsg)
            sys.exit(1)
        try:
            with open(ENCODINGS_FILE_PATH, "rb") as f:
                loaded_data = pickle.load(f)
            for key in ["encodings", "names", "thumbnails", "all_person_names"]:
                if key not in loaded_data:
                    raise KeyError(f"Missing key in encoding file: '{key}'")
            self.known_face_encodings      = loaded_data["encodings"]
            self.known_face_names          = loaded_data["names"]
            self.person_thumbnails_paths   = loaded_data["thumbnails"]
            self.all_known_person_names    = set(loaded_data["all_person_names"])
            print(f"  Loaded {len(self.all_known_person_names)} people.")
            if not self.known_face_encodings or not self.all_known_person_names:
                errmsg = f"FATAL: Encoding file is empty. Re-run encode_face_v2.py."
                print(errmsg)
                if QApplication.instance():
                    QMessageBox.critical(None, "Initialization Error", errmsg)
                sys.exit(1)
        except (pickle.UnpicklingError, EOFError, KeyError) as e:
            errmsg = f"FATAL: Corrupted/invalid encoding file:\n{e}\nRe-run encode_face_v2.py."
            print(errmsg)
            if QApplication.instance():
                QMessageBox.critical(None, "File Load Error", errmsg)
            sys.exit(1)
        except Exception as e:
            errmsg = f"FATAL: Unexpected error loading encodings:\n{e}"
            print(errmsg)
            traceback.print_exc()
            if QApplication.instance():
                QMessageBox.critical(None, "Load Error", errmsg)
            sys.exit(1)

    def _initialize_attendance(self):
        print(f"--- Resetting daily state for {self.today_date} ---")
        self.daily_attendance = {
            name: [("Absent", None)] * NUM_LECTURES
            for name in self.all_known_person_names
        }
        self.active_trackers.clear()
        self.tracker_details.clear()
        self.tracker_id_counter = 0
        for q in (identification_queue, results_queue):
            while not q.empty():
                try:
                    q.get_nowait()
                except queue.Empty:
                    break
        gc.collect()
        print("--- Daily state reset complete ---")

    def check_and_reset_daily_state(self):
        current_date = date.today()
        if current_date != self.today_date:
            print(f"*** Date changed: {self.today_date} -> {current_date} ***")
            self.save_daily_report(filename_date=self.today_date)
            self.today_date = current_date
            self._initialize_attendance()
            self.current_lecture = 1
            return True
        return False

    def set_lecture(self, lecture_num):
        if 1 <= lecture_num <= NUM_LECTURES:
            if self.current_lecture != lecture_num:
                print(f"--- Switched to Lecture {lecture_num} ---")
                self.current_lecture = lecture_num
        else:
            print(f"WARNING: Invalid lecture number: {lecture_num}")

    def start_identification_thread(self):
        self.identification_thread = threading.Thread(
            target=self._background_identifier, daemon=True
        )
        self.identification_thread.start()

    def _background_identifier(self):
        print("INFO: Identifier thread started.")
        while True:
            tracker_id = -1
            job = None
            try:
                job = identification_queue.get()
                if job is None:
                    break
                tracker_id, face_crop_rgb = job
                identified_name = UNKNOWN_FACE_LABEL

                if (face_crop_rgb is None or face_crop_rgb.size == 0 or
                        face_crop_rgb.shape[0] < 32 or face_crop_rgb.shape[1] < 32):
                    results_queue.put((tracker_id, UNKNOWN_FACE_LABEL))
                    identification_queue.task_done()
                    continue

                encodings = face_recognition.face_encodings(
                    face_crop_rgb, num_jitters=2, model="small"
                )
                if encodings and self.known_face_encodings:
                    distances = face_recognition.face_distance(
                        self.known_face_encodings, encodings[0]
                    )
                    if distances.size > 0:
                        best_idx = np.argmin(distances)
                        if distances[best_idx] < self.recognition_threshold:
                            identified_name = self.known_face_names[best_idx]

                results_queue.put((tracker_id, identified_name))
            except Exception as e:
                print(f"ID Thread ERROR: {type(e).__name__} - {e}")
                if tracker_id != -1:
                    results_queue.put((tracker_id, UNKNOWN_FACE_LABEL))
            finally:
                if job is not None:
                    identification_queue.task_done()
        print("INFO: Identifier thread finished.")

    def _get_bounding_box(self, tracker):
        pos = tracker.get_position()
        return int(pos.left()), int(pos.top()), int(pos.right()), int(pos.bottom())

    def _iou(self, boxA, boxB):
        xA = max(boxA[0], boxB[0]); yA = max(boxA[1], boxB[1])
        xB = min(boxA[2], boxB[2]); yB = min(boxA[3], boxB[3])
        inter = max(0, xB - xA) * max(0, yB - yA)
        denom = float(
            (boxA[2]-boxA[0])*(boxA[3]-boxA[1]) +
            (boxB[2]-boxB[0])*(boxB[3]-boxB[1]) - inter
        )
        return inter / denom if denom > 0 else 0.0

    def process_frame(self, frame_bgr):
        self.frame_count += 1
        current_time_obj = datetime.now()
        lecture_index  = self.current_lecture - 1
        frame_height, frame_width = frame_bgr.shape[:2]
        rgb_frame = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2RGB)

        # 1. Consume identification results
        try:
            while not results_queue.empty():
                tracker_id, identified_name = results_queue.get_nowait()
                if tracker_id in self.tracker_details:
                    self.tracker_details[tracker_id]['name'] = identified_name
                    if (identified_name != UNKNOWN_FACE_LABEL and
                            identified_name in self.daily_attendance and
                            0 <= lecture_index < NUM_LECTURES and
                            self.daily_attendance[identified_name][lecture_index][0] == "Absent"):
                        self.daily_attendance[identified_name][lecture_index] = (
                            "Present", current_time_obj
                        )
        except queue.Empty:
            pass
        except Exception as e:
            print(f"ERROR processing results_queue: {e}")

        trackers_to_remove = []

        # 2. Update existing trackers
        for tracker_id in list(self.active_trackers.keys()):
            if tracker_id not in self.tracker_details:
                continue
            tracker    = self.active_trackers[tracker_id]
            confidence = tracker.update(rgb_frame)

            if confidence >= TRACKER_CONFIDENCE_THRESHOLD:
                box  = self._get_bounding_box(tracker)
                det  = self.tracker_details[tracker_id]
                det['box']             = box
                det['last_seen_frame'] = self.frame_count
                name = det.get('name', UNKNOWN_FACE_LABEL)

                if (name != UNKNOWN_FACE_LABEL and name in self.daily_attendance and
                        0 <= lecture_index < NUM_LECTURES and
                        self.daily_attendance[name][lecture_index][0] == "Absent"):
                    self.daily_attendance[name][lecture_index] = ("Present", current_time_obj)

                # Emotion analysis
                can_analyze = False
                if self.deepface_functional and self.emotion_detection_active:
                    reconfirmed  = det.get('reconfirmed_this_cycle', False)
                    is_young     = (self.frame_count - det.get('initial_frame_count', self.frame_count)) < (FACE_DETECT_FRAME_SKIP * 2)
                    cur_emo      = det.get('emotion', DEFAULT_EMOTION_STATE)
                    if reconfirmed or is_young:
                        if (self.frame_count % EMOTION_ANALYZE_FRAME_SKIP == 0 or
                                cur_emo == DEFAULT_EMOTION_STATE):
                            can_analyze = True
                    elif cur_emo not in [DEFAULT_EMOTION_STATE, EMOTION_ERROR_STATE,
                                         EMOTION_OFF_STATE, EMOTION_DISABLED_STATE]:
                        det.update({'emotion': DEFAULT_EMOTION_STATE, 'emotion_details': None})

                if not self.deepface_functional:
                    det.update({'emotion': EMOTION_DISABLED_STATE, 'emotion_details': None})
                elif not self.emotion_detection_active:
                    det.update({'emotion': EMOTION_OFF_STATE, 'emotion_details': None})
                elif can_analyze:
                    try:
                        x1, y1, x2, y2 = box
                        if not (x1 < x2 and y1 < y2):
                            det.update({'emotion': EMOTION_ERROR_STATE, 'emotion_details': None})
                            continue
                        pad = 20
                        ry1 = max(0, y1 - pad); ry2 = min(frame_height, y2 + pad)
                        rx1 = max(0, x1 - pad); rx2 = min(frame_width,  x2 + pad)
                        if ry2 > ry1 and rx2 > rx1:
                            roi = frame_bgr[ry1:ry2, rx1:rx2]
                            results = DeepFace.analyze(
                                img_path=roi, actions=['emotion'],
                                detector_backend=EMOTION_DETECTOR_FOR_DEEPFACE,
                                enforce_detection=False, silent=True
                            )
                            if isinstance(results, list) and results:
                                r = results[0]
                                if 'dominant_emotion' in r and 'emotion' in r:
                                    det['emotion']         = r['dominant_emotion']
                                    det['emotion_details'] = r['emotion']
                                else:
                                    det.update({'emotion': EMOTION_ERROR_STATE, 'emotion_details': None})
                            else:
                                det.update({'emotion': EMOTION_ERROR_STATE, 'emotion_details': None})
                        else:
                            det.update({'emotion': EMOTION_ERROR_STATE, 'emotion_details': None})
                    except Exception:
                        det.update({'emotion': EMOTION_ERROR_STATE, 'emotion_details': None})
            else:
                if tracker_id not in trackers_to_remove:
                    trackers_to_remove.append(tracker_id)

        # 3. Periodic face detection
        if self.frame_count % FACE_DETECT_FRAME_SKIP == 0:
            for tid in self.active_trackers:
                if tid in self.tracker_details:
                    self.tracker_details[tid]['reconfirmed_this_cycle'] = False

            detected_boxes = []
            if (self.face_detect_model_to_use == "blazeface" and
                    self.blazeface_detector and BLAZEFACE_AVAILABLE):
                res = self.blazeface_detector.process(np.ascontiguousarray(rgb_frame))
                if res.detections:
                    for det in res.detections:
                        bb = det.location_data.relative_bounding_box
                        x1 = max(0, int(bb.xmin * frame_width))
                        y1 = max(0, int(bb.ymin * frame_height))
                        x2 = min(frame_width,  x1 + int(bb.width  * frame_width))
                        y2 = min(frame_height, y1 + int(bb.height * frame_height))
                        if x2 > x1 and y2 > y1:
                            detected_boxes.append((x1, y1, x2, y2))
            elif self.face_detect_model_to_use in ("hog", "cnn"):
                locs = face_recognition.face_locations(rgb_frame, model=self.face_detect_model_to_use)
                detected_boxes = [(l, t, r, b) for (t, r, b, l) in locs]

            assigned = [False] * len(detected_boxes)
            for tracker_id in list(self.active_trackers.keys()):
                if tracker_id not in self.tracker_details or 'box' not in self.tracker_details[tracker_id]:
                    continue
                tbox = self.tracker_details[tracker_id]['box']
                best_iou, best_j = 0.0, -1
                for j, dbox in enumerate(detected_boxes):
                    if not assigned[j]:
                        iou = self._iou(tbox, dbox)
                        if iou > 0.4 and iou > best_iou:
                            best_iou, best_j = iou, j
                if best_j != -1:
                    assigned[best_j] = True
                    self.tracker_details[tracker_id]['reconfirmed_this_cycle'] = True
                    self.tracker_details[tracker_id]['stale_cycles'] = 0
                else:
                    sc = self.tracker_details[tracker_id].get('stale_cycles', 0) + 1
                    self.tracker_details[tracker_id]['stale_cycles'] = sc
                    if sc > MAX_TRACKER_STALE_CYCLES and tracker_id not in trackers_to_remove:
                        trackers_to_remove.append(tracker_id)

            for i, dbox in enumerate(detected_boxes):
                if assigned[i]:
                    continue
                x1, y1, x2, y2 = dbox
                if not (0 <= x1 < x2 <= frame_width and 0 <= y1 < y2 <= frame_height):
                    continue
                tracker = dlib.correlation_tracker()
                tracker.start_track(rgb_frame, dlib.rectangle(x1, y1, x2, y2))
                nid = self.tracker_id_counter
                self.tracker_id_counter += 1
                self.active_trackers[nid] = tracker
                init_emo = EMOTION_DISABLED_STATE
                if self.deepface_functional:
                    init_emo = EMOTION_OFF_STATE if not self.emotion_detection_active else DEFAULT_EMOTION_STATE
                self.tracker_details[nid] = {
                    'name': UNKNOWN_FACE_LABEL, 'box': dbox,
                    'last_seen_frame': self.frame_count,
                    'initial_frame_count': self.frame_count,
                    'emotion': init_emo, 'emotion_details': None,
                    'reconfirmed_this_cycle': True, 'stale_cycles': 0
                }
                pad_enc = 10
                cy1 = max(0, y1 - pad_enc); cy2 = min(frame_height, y2 + pad_enc)
                cx1 = max(0, x1 - pad_enc); cx2 = min(frame_width,  x2 + pad_enc)
                if cy1 < cy2 and cx1 < cx2:
                    crop = rgb_frame[cy1:cy2, cx1:cx2].copy()
                    if crop.size > 0 and not identification_queue.full():
                        identification_queue.put((nid, crop))

        # 4. Remove stale trackers
        for tid in set(trackers_to_remove):
            self.active_trackers.pop(tid, None)
            self.tracker_details.pop(tid, None)

        # 5. Prepare draw data for UI
        boxes_out, names_out, emotions_out, details_out = [], [], [], {}
        threshold_frames = FACE_DETECT_FRAME_SKIP * (MAX_TRACKER_STALE_CYCLES + 1)
        for tid, det in list(self.tracker_details.items()):
            age = self.frame_count - det.get('last_seen_frame', self.frame_count + threshold_frames + 1)
            if age < threshold_frames and 'box' in det:
                boxes_out.append(det['box'])
                names_out.append(det.get('name', UNKNOWN_FACE_LABEL))
                emo = det.get('emotion', EMOTION_ERROR_STATE)
                if not self.deepface_functional:
                    emo = EMOTION_DISABLED_STATE
                elif not self.emotion_detection_active and emo != EMOTION_DISABLED_STATE:
                    emo = EMOTION_OFF_STATE
                emotions_out.append(emo)
                details_out[tid] = det.copy()

        return boxes_out, names_out, emotions_out, details_out

    def set_emotion_details_display_status(self, should_display: bool):
        self.display_emotion_details_globally = should_display

    def set_emotion_detection_active(self, active: bool):
        if not self.deepface_functional:
            self.emotion_detection_active = False
            self.display_emotion_details_globally = False
            for tid in list(self.tracker_details.keys()):
                if tid in self.tracker_details:
                    self.tracker_details[tid].update({'emotion': EMOTION_DISABLED_STATE, 'emotion_details': None})
            return
        self.emotion_detection_active = active
        if not active:
            self.display_emotion_details_globally = False
            for tid in list(self.tracker_details.keys()):
                if tid in self.tracker_details:
                    det = self.tracker_details[tid]
                    if det.get('emotion') != EMOTION_DISABLED_STATE:
                        det.update({'emotion': EMOTION_OFF_STATE, 'emotion_details': None})
        else:
            for tid in list(self.tracker_details.keys()):
                if tid in self.tracker_details:
                    det = self.tracker_details[tid]
                    if det.get('emotion') in (EMOTION_OFF_STATE, EMOTION_DISABLED_STATE):
                        det['emotion'] = DEFAULT_EMOTION_STATE

    def save_daily_report(self, filename_date=None):
        if not OPENPYXL_AVAILABLE:
            return False, "Cannot save: openpyxl not available. Install with: pip install openpyxl"
        report_date     = filename_date or self.today_date
        report_date_str = report_date.strftime("%Y-%m-%d")
        filename  = f"Daily_Attendance_{report_date_str}.xlsx"
        filepath  = os.path.join(ATTENDANCE_DAY_DIR, filename)
        print(f"--- Saving report to '{filepath}' ---")
        try:
            wb    = Workbook()
            sheet = wb.active
            sheet.title = f"Attendance {report_date_str}"
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header = ["Student Name"]
            for i in range(1, NUM_LECTURES + 1):
                header.extend([f"L{i} Status", f"L{i} Time"])
            sheet.append(header)
            for name in sorted(self.all_known_person_names):
                records  = self.daily_attendance.get(name, [("Absent", None)] * NUM_LECTURES)
                row_data = [name]
                for status, ts in records:
                    row_data.append(status)
                    row_data.append(ts.strftime('%H:%M:%S') if isinstance(ts, datetime) else "--:--:--")
                sheet.append(row_data)
            for cell in sheet[1]:
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_idx, col_cells in enumerate(sheet.columns, 1):
                max_len = 0
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[get_column_letter(col_idx)].width = min((max_len + 2) * 1.1, 50)
            wb.save(filepath)
            print(f"Report saved: '{filename}'")
            return True, f"Report saved to:\n{filepath}"
        except PermissionError:
            msg = f"Permission denied for '{filepath}'. Is it open in Excel?"
            print(f"ERROR: {msg}")
            return False, msg
        except Exception as e:
            msg = f"Error saving '{filepath}': {type(e).__name__} - {e}"
            print(f"ERROR: {msg}")
            traceback.print_exc(limit=1)
            return False, msg

    def stop_identifier_thread(self):
        try:
            identification_queue.put(None, block=True, timeout=1.0)
        except queue.Full:
            print("WARNING: ID queue full, stop signal not sent in time.")
        except Exception as e:
            print(f"ERROR sending stop signal: {e}")

    def close(self):
        print("INFO: Closing system resources...")
        self.set_emotion_detection_active(False)
        self.stop_identifier_thread()
        if hasattr(self, 'identification_thread') and self.identification_thread.is_alive():
            self.identification_thread.join(timeout=2.0)
        if self.blazeface_detector and BLAZEFACE_AVAILABLE:
            try:
                self.blazeface_detector.close()
            except Exception as e:
                print(f"WARNING: Error closing BlazeFace: {e}")
        self.blazeface_detector = None
        print("INFO: System cleanup done.")


# ==============================================================================
# Emotion Detail Widget
# ==============================================================================
class EmotionDetailWidget(QWidget):
    window_counter = 0

    def __init__(self, tracker_id, person_name, emotion_data, dominant_emotion, parent=None):
        super().__init__(parent)
        self.tracker_id      = tracker_id
        self.person_name     = person_name
        self.emotion_data    = emotion_data or {}
        self.dominant_emotion = dominant_emotion

        self.setWindowFlags(
            Qt.WindowType.Window | Qt.WindowType.CustomizeWindowHint |
            Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowStaysOnTopHint
        )
        self.setWindowTitle(f"Emotion: {person_name} (ID: {tracker_id})")
        self.setFixedSize(DETAIL_WINDOW_WIDTH_QT, DETAIL_WINDOW_HEIGHT_QT)
        self.setStyleSheet(f"background-color: {DETAIL_BG_COLOR_QT.name()};")
        self._position_window(parent)
        EmotionDetailWidget.window_counter += 1

    def _position_window(self, main_win):
        screen  = QApplication.primaryScreen().availableGeometry()
        base_x  = screen.right() - DETAIL_WINDOW_WIDTH_QT - 20
        base_y  = screen.top() + 50
        if main_win and main_win.isVisible():
            mg = main_win.geometry()
            if mg.right() + 10 + DETAIL_WINDOW_WIDTH_QT <= screen.right():
                base_x, base_y = mg.right() + 10, mg.top() + 20
            elif mg.left() - DETAIL_WINDOW_WIDTH_QT - 10 >= screen.left():
                base_x, base_y = mg.left() - DETAIL_WINDOW_WIDTH_QT - 10, mg.top() + 20
        offset_x = (EmotionDetailWidget.window_counter % 3) * 30
        offset_y = (EmotionDetailWidget.window_counter // 3) * (DETAIL_WINDOW_HEIGHT_QT // 3)
        fx = max(screen.left(), min(base_x - offset_x, screen.right() - DETAIL_WINDOW_WIDTH_QT))
        fy = max(screen.top(),  min(base_y + offset_y, screen.bottom() - DETAIL_WINDOW_HEIGHT_QT))
        self.move(fx, fy)

    def paintEvent(self, event):
        painter  = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        cy = 15; lh = 18; lm = 15
        bar_max_w = DETAIL_WINDOW_WIDTH_QT - lm * 2 - 70
        bar_h = 10

        painter.setPen(DETAIL_HEADER_COLOR_QT)
        painter.setFont(QFont(DETAIL_FONT_FAMILY, 10, QFont.Weight.Bold))
        painter.drawText(QRect(lm, cy, self.width() - 2*lm, lh+4),
                         Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                         f"Subject: {self.person_name}")
        cy += lh + 4 + 6

        dominant_score = self.emotion_data.get(self.dominant_emotion, 0.0)
        painter.setPen(DETAIL_DOMINANT_COLOR_QT)
        painter.setFont(QFont(DETAIL_FONT_FAMILY, 9, QFont.Weight.DemiBold))
        painter.drawText(QRect(lm, cy, self.width() - 2*lm, lh),
                         Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                         f"Dominant: {self.dominant_emotion.capitalize()} ({dominant_score:.1f}%)")
        cy += lh + 6

        painter.setPen(DETAIL_PERCENT_COLOR_QT)
        painter.setFont(QFont(DETAIL_FONT_FAMILY, 9, QFont.Weight.Normal))
        painter.drawText(QRect(lm, cy, self.width() - 2*lm, lh),
                         Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                         "Emotion Profile:")
        cy += int(lh * 1.1)

        painter.setFont(QFont(DETAIL_FONT_FAMILY, 8))
        label_w = 65; pct_w = 35; bar_x = lm + label_w
        for emo in EMOTION_LABELS:
            score = self.emotion_data.get(emo, 0.0)
            painter.setPen(DETAIL_PERCENT_COLOR_QT)
            painter.drawText(QRect(lm, cy, label_w - 5, lh),
                             Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
                             f"{emo.capitalize()}:")
            bar_y = cy + (lh - bar_h) // 2
            painter.setPen(DETAIL_BAR_OUTLINE_COLOR_QT)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(bar_x, bar_y, bar_max_w, bar_h)
            bar_w = max(0, min(int((score / 100.0) * (bar_max_w - 2)), bar_max_w - 2))
            bar_color = DETAIL_DOMINANT_COLOR_QT if emo == self.dominant_emotion else DETAIL_BAR_COLOR_QT
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(bar_color))
            if bar_w > 0:
                painter.drawRect(bar_x + 1, bar_y + 1, bar_w, bar_h - 2)
            painter.setPen(DETAIL_PERCENT_COLOR_QT)
            painter.drawText(QRect(bar_x + bar_max_w + 5, cy, pct_w, lh),
                             Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                             f"{score:.0f}%")
            cy += lh

    def update_data(self, person_name, emotion_data, dominant_emotion):
        self.person_name      = person_name
        self.emotion_data     = emotion_data or {}
        self.dominant_emotion = dominant_emotion
        self.setWindowTitle(f"Emotion: {person_name} (ID: {self.tracker_id})")
        self.update()

    def closeEvent(self, event):
        EmotionDetailWidget.window_counter = max(0, EmotionDetailWidget.window_counter - 1)
        super().closeEvent(event)

    @staticmethod
    def reset_window_counter():
        EmotionDetailWidget.window_counter = 0


# ==============================================================================
# Main Application Window
# ==============================================================================
class App(QMainWindow):
    def __init__(self, known_faces_dir, video_source=LAPTOP_CAM_SRC):
        super().__init__()
        self.setWindowTitle(WINDOW_TITLE_QT)
        self.setGeometry(50, 50, 1500, 850)
        self.setMinimumSize(1200, 700)

        self.initial_video_source    = video_source
        self.active_camera_source_id = None
        self.vid_capture             = None
        self.camera_worker           = None
        self.switching_camera_lock   = threading.Lock()

        self.frame_width  = 0
        self.frame_height = 0
        self.ui_update_delay_ms = 33

        self.emotion_windows_visibility_globally_enabled = False
        self.emotion_detection_globally_active           = False
        self.active_emotion_detail_windows               = {}
        self.closing_app                                 = False

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame_and_ui)

        self.recognition_system = None
        try:
            print("--- Initializing Backend ---")
            self.recognition_system = IntegratedAttendanceEmotionSystem(known_faces_dir)
            print("--- Backend Ready ---")
        except SystemExit:
            raise
        except Exception as e:
            errmsg = f"CRITICAL ERROR during backend init:\n{type(e).__name__}: {e}"
            print(errmsg)
            traceback.print_exc()
            if QApplication.instance():
                QMessageBox.critical(None, "Fatal Initialization Error",
                                     errmsg + "\nApplication will close.")
            raise RuntimeError("Backend Init Failed") from e

        self.setup_ui()
        self._update_emotion_buttons_state()
        print("--- Setting up camera ---")
        self.attempt_camera_start(self.initial_video_source)
        print("--- Application ready ---")

    def setup_ui(self):
        self.setStyleSheet(self._get_stylesheet())
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)

        # Video panel
        video_panel = QWidget()
        video_panel.setObjectName("VideoPanel")
        video_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        vpl = QVBoxLayout(video_panel)
        self.video_label = QLabel("Initializing Camera...")
        self.video_label.setObjectName("VideoLabel")
        self.video_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.video_label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Ignored)
        vpl.addWidget(self.video_label)
        main_layout.addWidget(video_panel, 7)

        # Right panel
        right_panel = QWidget()
        right_panel.setObjectName("RightPanel")
        rpl = QVBoxLayout(right_panel)

        # Top controls
        top_ctrl = QFrame()
        top_ctrl.setObjectName("ControlsFrame")
        tcl = QGridLayout(top_ctrl)

        tcl.addWidget(QLabel("Lecture:"), 0, 0, Qt.AlignmentFlag.AlignRight)
        lbg = QWidget()
        lblay = QHBoxLayout(lbg)
        lblay.setContentsMargins(0, 0, 0, 0)
        lblay.setSpacing(5)
        self.lecture_buttons = []
        for i in range(1, NUM_LECTURES + 1):
            btn = QPushButton(str(i))
            btn.setCheckable(True)
            btn.setAutoExclusive(True)
            btn.clicked.connect(lambda checked, l=i: self.select_lecture(l) if checked else None)
            lblay.addWidget(btn)
            self.lecture_buttons.append(btn)
        tcl.addWidget(lbg, 0, 1)
        self.selected_lecture_label = QLabel("Lec 1 ACTIVE")
        self.selected_lecture_label.setObjectName("StatusLabel")
        tcl.addWidget(self.selected_lecture_label, 0, 2, Qt.AlignmentFlag.AlignLeft)

        tcl.addWidget(QLabel("Camera:"), 1, 0, Qt.AlignmentFlag.AlignRight)
        self.btn_laptop_cam = QPushButton("Laptop Cam")
        self.btn_laptop_cam.setCheckable(True)
        self.btn_laptop_cam.clicked.connect(self.use_laptop_camera)
        self.btn_iphone_cam = QPushButton("Phone Cam")
        self.btn_iphone_cam.setCheckable(True)
        self.btn_iphone_cam.clicked.connect(self.use_iphone_camera)
        cam_lay = QHBoxLayout()
        cam_lay.addWidget(self.btn_laptop_cam)
        cam_lay.addWidget(self.btn_iphone_cam)
        tcl.addLayout(cam_lay, 1, 1, 1, 2)

        tcl.addWidget(QLabel("Emotion AI:"), 2, 0, Qt.AlignmentFlag.AlignRight)
        self.btn_start_emotion = QPushButton("Start Emo")
        self.btn_start_emotion.setObjectName("btn_start_emotion")
        self.btn_start_emotion.clicked.connect(self.start_emotion_detection_ui)
        self.btn_stop_emotion = QPushButton("Stop Emo")
        self.btn_stop_emotion.setObjectName("btn_stop_emotion")
        self.btn_stop_emotion.clicked.connect(self.stop_emotion_detection_ui)
        self.btn_toggle_emotion_visibility = QPushButton("Show Details")
        self.btn_toggle_emotion_visibility.setObjectName("btn_toggle_emotion_visibility")
        self.btn_toggle_emotion_visibility.clicked.connect(self.toggle_emotion_windows_visibility_ui)
        emo_lay = QHBoxLayout()
        emo_lay.addWidget(self.btn_start_emotion)
        emo_lay.addWidget(self.btn_stop_emotion)
        emo_lay.addWidget(self.btn_toggle_emotion_visibility)
        tcl.addLayout(emo_lay, 2, 1, 1, 2)
        tcl.setColumnStretch(1, 1)
        rpl.addWidget(top_ctrl)

        # Attendance table
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(5)
        self.attendance_table.setHorizontalHeaderLabels(["#", "Pic", "Name", "Status", "Time"])
        hh = self.attendance_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self.attendance_table.setColumnWidth(0, 30)
        self.attendance_table.setColumnWidth(1, THUMBNAIL_SIZE_QT.width() + 10)
        self.attendance_table.setColumnWidth(3, 80)
        self.attendance_table.setColumnWidth(4, 70)
        self.attendance_table.verticalHeader().setVisible(False)
        self.attendance_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.attendance_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.attendance_table.setAlternatingRowColors(True)
        rpl.addWidget(self.attendance_table)

        # Bottom controls
        bot_ctrl = QFrame()
        bot_ctrl.setObjectName("ControlsFrame")
        bcl = QHBoxLayout(bot_ctrl)
        self.btn_reset = QPushButton("Reset Today's Log")
        self.btn_reset.setObjectName("ResetButton")
        self.btn_reset.clicked.connect(self.reset_attendance)
        self.btn_save = QPushButton("Save Daily Report")
        self.btn_save.setObjectName("SaveButton")
        self.btn_save.clicked.connect(self.save_report)
        bcl.addStretch()
        bcl.addWidget(self.btn_reset)
        bcl.addWidget(self.btn_save)
        bcl.addStretch()
        rpl.addWidget(bot_ctrl)

        main_layout.addWidget(right_panel, 3)

        if self.recognition_system:
            self.select_lecture(self.recognition_system.current_lecture)
            self.populate_attendance_table_structure()
        self._update_camera_buttons_state()

    def _get_stylesheet(self):
        return """
            QMainWindow, QWidget { background-color: #2E2E3A; }
            QWidget#VideoPanel { background-color: #1C1C1C; border-radius: 5px; }
            QLabel#VideoLabel { background-color: #101010; color: #D0D0D0; font-size: 11pt; }
            QWidget#RightPanel { background-color: #353542; border-radius: 5px; }
            QFrame#ControlsFrame { background-color: transparent; padding: 5px; margin-bottom: 3px; }
            QLabel { color: #E0E0E0; font-size: 10pt; font-family: "Segoe UI", Arial, sans-serif; }
            QLabel#StatusLabel { color: #87CEEB; font-weight: bold; padding-left: 10px; }
            QPushButton {
                background-color: #4A4A5A; color: #F0F0F0; border: 1px solid #5A5A6A;
                padding: 8px 12px; border-radius: 4px; font-size: 9pt; min-height: 20px;
            }
            QPushButton:hover { background-color: #5A5A6A; }
            QPushButton:pressed { background-color: #3A3A4A; }
            QPushButton:disabled { background-color: #404048; color: #888888; border-color: #505055; }
            QPushButton:checked { background-color: #0078D7; color: white; border: 1px solid #005A9E; font-weight: bold; }
            QPushButton#SaveButton { background-color: #28A745; color: white; }
            QPushButton#SaveButton:hover { background-color: #218838; }
            QPushButton#ResetButton { background-color: #DC3545; color: white; }
            QPushButton#ResetButton:hover { background-color: #C82333; }
            QPushButton#btn_start_emotion[active_state="true"] { background-color: #28A745; color: white; font-weight: bold; }
            QPushButton#btn_stop_emotion[active_state="true"]  { background-color: #DC3545; color: white; font-weight: bold; }
            QPushButton#btn_toggle_emotion_visibility[showing="true"] { background-color: #17A2B8; color: white; font-weight: bold; }
            QTableWidget {
                background-color: #3C3C4A; color: #E0E0E0; gridline-color: #4A4A5A;
                border: 1px solid #4A4A5A; font-size: 9pt; border-radius: 4px;
            }
            QTableWidget::item { padding: 4px; }
            QTableWidget::item:selected { background-color: #0078D7; color: white; }
            QTableWidget::item:alternate { background-color: #424250; }
            QHeaderView::section {
                background-color: #4A4A5A; color: #F0F0F0; padding: 5px;
                border: 1px solid #5A5A6A; font-weight: bold;
            }
            QScrollBar:vertical { border: none; background: #3C3C4A; width: 12px; margin: 0px; }
            QScrollBar::handle:vertical { background: #5A5A6A; min-height: 20px; border-radius: 6px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; background: transparent; }
            QTableWidget QWidget { background-color: transparent; color: #E0E0E0; }
        """

    def _update_emotion_buttons_state(self):
        if self.closing_app or not self.recognition_system:
            return
        self.btn_start_emotion.setProperty("active_state", "false")
        self.btn_stop_emotion.setProperty("active_state", "false")
        self.btn_toggle_emotion_visibility.setProperty("showing", "false")

        df_ok = self.recognition_system.deepface_functional
        if not df_ok:
            self.btn_start_emotion.setEnabled(False);  self.btn_start_emotion.setText("Emo N/A")
            self.btn_stop_emotion.setEnabled(False);   self.btn_stop_emotion.setText("Stop Emo")
            self.btn_toggle_emotion_visibility.setEnabled(False)
            self.btn_toggle_emotion_visibility.setText("Details N/A")
        else:
            self.btn_start_emotion.setText("Start Emo")
            self.btn_stop_emotion.setText("Stop Emo")
            self.btn_toggle_emotion_visibility.setText("Show Details")
            if self.emotion_detection_globally_active:
                self.btn_start_emotion.setEnabled(False)
                self.btn_start_emotion.setProperty("active_state", "true")
                self.btn_stop_emotion.setEnabled(True)
                self.btn_toggle_emotion_visibility.setEnabled(True)
                if self.emotion_windows_visibility_globally_enabled:
                    self.btn_toggle_emotion_visibility.setText("Hide Details")
                    self.btn_toggle_emotion_visibility.setProperty("showing", "true")
            else:
                self.btn_start_emotion.setEnabled(True)
                self.btn_stop_emotion.setEnabled(False)
                self.btn_stop_emotion.setProperty("active_state", "true")
                self.btn_toggle_emotion_visibility.setEnabled(False)

        for btn in (self.btn_start_emotion, self.btn_stop_emotion, self.btn_toggle_emotion_visibility):
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def start_emotion_detection_ui(self):
        if not self.recognition_system or not self.recognition_system.deepface_functional:
            QMessageBox.warning(self, "Emotion AI Unavailable",
                                "DeepFace is not functional or available.")
            return
        self.emotion_detection_globally_active = True
        self.recognition_system.set_emotion_detection_active(True)
        self.emotion_windows_visibility_globally_enabled = False
        self.recognition_system.set_emotion_details_display_status(False)
        self._close_all_emotion_detail_windows()
        self._update_emotion_buttons_state()

    def stop_emotion_detection_ui(self):
        if not self.recognition_system:
            return
        self.emotion_detection_globally_active = False
        self.recognition_system.set_emotion_detection_active(False)
        self.emotion_windows_visibility_globally_enabled = False
        self.recognition_system.set_emotion_details_display_status(False)
        self._close_all_emotion_detail_windows()
        self._update_emotion_buttons_state()

    def toggle_emotion_windows_visibility_ui(self):
        if (not self.recognition_system or
                not self.recognition_system.deepface_functional or
                not self.emotion_detection_globally_active):
            QMessageBox.warning(self, "Cannot Toggle Details",
                                "Emotion detection must be active first.")
            return
        self.emotion_windows_visibility_globally_enabled = not self.emotion_windows_visibility_globally_enabled
        self.recognition_system.set_emotion_details_display_status(
            self.emotion_windows_visibility_globally_enabled
        )
        self._update_emotion_buttons_state()
        if not self.emotion_windows_visibility_globally_enabled:
            self._close_all_emotion_detail_windows()

    def _manage_emotion_detail_windows(self, tracker_details_from_backend):
        if self.closing_app or not self.recognition_system:
            return
        current_ids = set(tracker_details_from_backend.keys())
        ui_ids      = set(self.active_emotion_detail_windows.keys())
        for tid in ui_ids - current_ids:
            self._destroy_emotion_detail_window(tid)
        if (not self.emotion_windows_visibility_globally_enabled or
                not self.emotion_detection_globally_active or
                not self.recognition_system.deepface_functional):
            if self.active_emotion_detail_windows:
                self._close_all_emotion_detail_windows()
            return
        if not self.active_emotion_detail_windows:
            EmotionDetailWidget.reset_window_counter()
        for tid, det in tracker_details_from_backend.items():
            name = det.get('name', UNKNOWN_FACE_LABEL)
            emo_data  = det.get('emotion_details')
            dominant  = det.get('emotion', EMOTION_ERROR_STATE)
            should_show = (
                name != UNKNOWN_FACE_LABEL and
                isinstance(emo_data, dict) and emo_data and
                dominant not in (EMOTION_ERROR_STATE, EMOTION_DISABLED_STATE,
                                 EMOTION_OFF_STATE, DEFAULT_EMOTION_STATE, "...")
            )
            if should_show:
                if tid not in self.active_emotion_detail_windows:
                    win = EmotionDetailWidget(tid, name, emo_data, dominant, parent=self)
                    self.active_emotion_detail_windows[tid] = win
                    win.show()
                else:
                    self.active_emotion_detail_windows[tid].update_data(name, emo_data, dominant)
            else:
                self._destroy_emotion_detail_window(tid)

    def _destroy_emotion_detail_window(self, tracker_id):
        win = self.active_emotion_detail_windows.pop(tracker_id, None)
        if win:
            win.close()
            win.deleteLater()

    def _close_all_emotion_detail_windows(self):
        for tid in list(self.active_emotion_detail_windows.keys()):
            self._destroy_emotion_detail_window(tid)
        self.active_emotion_detail_windows.clear()
        EmotionDetailWidget.reset_window_counter()

    def attempt_camera_start(self, source_id):
        if not self.switching_camera_lock.acquire(blocking=False):
            return
        try:
            if (self.active_camera_source_id == source_id and
                    self.vid_capture and self.vid_capture.isOpened()):
                self._update_camera_buttons_state()
                if hasattr(self, 'timer') and not self.timer.isActive():
                    self.timer.start(self.ui_update_delay_ms)
                return

            if hasattr(self, 'timer') and self.timer.isActive():
                self.timer.stop()
            if self.camera_worker and self.camera_worker.isRunning():
                self.camera_worker.stop()
            self.camera_worker = None
            if self.vid_capture:
                self.vid_capture.release()
            self.vid_capture = None
            self.active_camera_source_id = None
            self.video_label.setText(f"Connecting to: {str(source_id)[:30]}...")
            self.video_label.setStyleSheet("color: yellow; font-weight: bold;")
            QApplication.processEvents()

            self.camera_worker = CameraWorker(source_id, REQUEST_RESOLUTION)
            self.camera_worker.camera_opened_signal.connect(self._on_camera_opened)
            self.camera_worker.camera_error_signal.connect(self._on_camera_error)
            self.camera_worker.finished.connect(self._on_camera_worker_finished)
            self.camera_worker.start()
        finally:
            self.switching_camera_lock.release()

    def _on_camera_opened(self, success, cap_object, width, height):
        if self.closing_app:
            if cap_object:
                cap_object.release()
            return
        if success and cap_object:
            self.vid_capture  = cap_object
            self.frame_width  = width
            self.frame_height = height
            self.active_camera_source_id = self.camera_worker.source if self.camera_worker else "Unknown"
            self.video_label.setText("")
            if not self.timer.isActive():
                self.timer.start(self.ui_update_delay_ms)
        else:
            self.vid_capture = None
            self.active_camera_source_id = None
            if not self.closing_app:
                src = self.camera_worker.source if self.camera_worker else "N/A"
                self._show_video_error_message(f"Failed to open camera: {str(src)[:30]}")
        self._update_camera_buttons_state()

    def _on_camera_error(self, message):
        if self.closing_app:
            return
        self.vid_capture = None
        self.active_camera_source_id = None
        self._show_video_error_message(message)
        self._update_camera_buttons_state()
        if self.timer.isActive():
            self.timer.stop()

    def _on_camera_worker_finished(self):
        sender = self.sender()
        if sender:
            sender.deleteLater()
            if self.camera_worker == sender:
                self.camera_worker = None
        if not (self.vid_capture and self.vid_capture.isOpened()):
            self.active_camera_source_id = None
            if self.timer.isActive():
                self.timer.stop()
        self._update_camera_buttons_state()

    def use_laptop_camera(self):
        if (self.active_camera_source_id != LAPTOP_CAM_SRC or
                not (self.vid_capture and self.vid_capture.isOpened())):
            self.btn_iphone_cam.setChecked(False)
            self.attempt_camera_start(LAPTOP_CAM_SRC)

    def use_iphone_camera(self):
        if (self.active_camera_source_id != IPHONE_CAM_SRC or
                not (self.vid_capture and self.vid_capture.isOpened())):
            self.btn_laptop_cam.setChecked(False)
            self.attempt_camera_start(IPHONE_CAM_SRC)

    def _update_camera_buttons_state(self):
        if self.closing_app:
            return
        laptop_active = (self.active_camera_source_id == LAPTOP_CAM_SRC and
                         self.vid_capture and self.vid_capture.isOpened())
        phone_active  = (self.active_camera_source_id == IPHONE_CAM_SRC and
                         self.vid_capture and self.vid_capture.isOpened())
        self.btn_laptop_cam.setChecked(laptop_active)
        self.btn_iphone_cam.setChecked(phone_active)
        self.btn_laptop_cam.setEnabled(not laptop_active)
        self.btn_iphone_cam.setEnabled(not phone_active)

    def _show_video_error_message(self, message):
        print(f"UI VIDEO ERROR: {message}")
        self.video_label.setText(f"Video Error:\n{str(message)[:100]}\n\nCheck camera and try again.")
        self.video_label.setStyleSheet(
            "color: red; font-weight: bold; font-size: 12pt; qproperty-alignment: AlignCenter;"
        )

    def select_lecture(self, lecture_num):
        if self.recognition_system:
            self.recognition_system.set_lecture(lecture_num)
        self.selected_lecture_label.setText(f"Lec {lecture_num} ACTIVE")
        for i, btn in enumerate(self.lecture_buttons):
            btn.setChecked((i + 1) == lecture_num)
        self.update_attendance_table_display()

    def save_report(self):
        if not self.recognition_system:
            return
        success, message = self.recognition_system.save_daily_report()
        if success:
            QMessageBox.information(self, "Report Saved", message)
        else:
            QMessageBox.critical(self, "Save Error", message)

    def reset_attendance(self):
        if not self.recognition_system:
            return
        reply = QMessageBox.question(
            self, "Confirm Reset",
            "Reset ALL attendance data for today?\nThis cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.recognition_system._initialize_attendance()
            self.populate_attendance_table_structure()
            self.update_attendance_table_display()
            self.select_lecture(1)
            QMessageBox.information(self, "Reset Complete", "Today's attendance log has been reset.")

    def populate_attendance_table_structure(self):
        self.attendance_table.setRowCount(0)
        if not self.recognition_system or not self.recognition_system.all_known_person_names:
            return
        sorted_names = sorted(self.recognition_system.all_known_person_names)
        self.attendance_table.setRowCount(len(sorted_names))
        for i, name in enumerate(sorted_names):
            num_item = QTableWidgetItem(str(i + 1))
            num_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.attendance_table.setItem(i, 0, num_item)

            thumb_container = QWidget()
            tl = QHBoxLayout(thumb_container)
            tl.setContentsMargins(2, 2, 2, 2)
            tl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            thumb_label = QLabel()
            thumb_label.setFixedSize(THUMBNAIL_SIZE_QT)
            thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            path = self.recognition_system.person_thumbnails_paths.get(name)
            if path and os.path.exists(path):
                px = QPixmap(path)
                if not px.isNull():
                    thumb_label.setPixmap(px.scaled(
                        THUMBNAIL_SIZE_QT,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation
                    ))
                else:
                    thumb_label.setText("Err")
            else:
                thumb_label.setText("N/P")
            tl.addWidget(thumb_label)
            self.attendance_table.setCellWidget(i, 1, thumb_container)
            self.attendance_table.setRowHeight(i, THUMBNAIL_SIZE_QT.height() + 6)

            self.attendance_table.setItem(i, 2, QTableWidgetItem(name))
            si = QTableWidgetItem("Loading...")
            si.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            ti = QTableWidgetItem("--:--:--")
            ti.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.attendance_table.setItem(i, 3, si)
            self.attendance_table.setItem(i, 4, ti)
        self.update_attendance_table_display()

    def update_attendance_table_display(self):
        if self.closing_app or not hasattr(self, 'attendance_table') or not self.recognition_system:
            return
        lec_idx = self.recognition_system.current_lecture - 1
        present_bg = QColor("#C6EFCE"); present_fg = QColor("#006100")
        absent_bg  = QColor("#FFC7CE"); absent_fg  = QColor("#9C0006")
        error_bg   = QColor("orange");  error_fg   = QColor("black")

        for row in range(self.attendance_table.rowCount()):
            name_item = self.attendance_table.item(row, 2)
            if not name_item:
                continue
            name       = name_item.text()
            status_item = self.attendance_table.item(row, 3)
            time_item   = self.attendance_table.item(row, 4)
            if not status_item or not time_item:
                continue

            status_val = "Error"; time_str = "??:??"; bg = error_bg; fg = error_fg
            if name in self.recognition_system.daily_attendance:
                records = self.recognition_system.daily_attendance[name]
                if 0 <= lec_idx < len(records):
                    status_val, ts = records[lec_idx]
                    if status_val == "Present":
                        time_str = ts.strftime('%H:%M:%S') if isinstance(ts, datetime) else "Present"
                        bg, fg = present_bg, present_fg
                    else:
                        time_str = "--:--:--"
                        bg, fg = absent_bg, absent_fg
                else:
                    status_val, time_str = "Lec Err", "??:??"
            else:
                status_val, time_str = "No Data", "??:??"

            status_item.setText(status_val); status_item.setBackground(bg); status_item.setForeground(fg)
            time_item.setText(time_str);     time_item.setBackground(bg);   time_item.setForeground(fg)

    def update_frame_and_ui(self):
        if (self.closing_app or self.switching_camera_lock.locked() or
                not self.recognition_system or not self.vid_capture or
                not self.vid_capture.isOpened()):
            if (self.active_camera_source_id is not None and
                    not (self.vid_capture and self.vid_capture.isOpened()) and
                    not self.switching_camera_lock.locked() and not self.closing_app):
                self._show_video_error_message(
                    f"Camera {str(self.active_camera_source_id)[:30]} disconnected."
                )
                self.active_camera_source_id = None
                self._update_camera_buttons_state()
                if self.timer.isActive():
                    self.timer.stop()
            return

        if self.recognition_system.check_and_reset_daily_state():
            self.populate_attendance_table_structure()
            self.select_lecture(self.recognition_system.current_lecture)

        ret, frame_bgr = self.vid_capture.read()
        if not ret or frame_bgr is None:
            if not self.closing_app:
                self._show_video_error_message(
                    f"Failed to read frame from {str(self.active_camera_source_id)[:30]}."
                )
                self.active_camera_source_id = None
                self._update_camera_buttons_state()
                if self.timer.isActive():
                    self.timer.stop()
            return

        try:
            boxes, names, emotions, tracker_details = self.recognition_system.process_frame(frame_bgr)
            draw_frame = frame_bgr.copy()
            unknown_clr = (0, 0, 255); known_clr = (0, 255, 0); text_clr = (255, 255, 255)
            font_scale = 0.45; thickness = 1

            for box, name, emotion_str in zip(boxes, names, emotions):
                x1, y1, x2, y2 = map(int, box)
                box_color = unknown_clr if name == UNKNOWN_FACE_LABEL else known_clr
                cv2.rectangle(draw_frame, (x1, y1), (x2, y2), box_color, 2)

                display_text = name
                if (self.recognition_system.deepface_functional and
                        self.emotion_detection_globally_active and
                        emotion_str not in (EMOTION_OFF_STATE, EMOTION_DISABLED_STATE,
                                            DEFAULT_EMOTION_STATE, EMOTION_ERROR_STATE, "...")):
                    display_text += f" ({emotion_str.capitalize()})"

                (tw, th), baseline = cv2.getTextSize(
                    display_text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness
                )
                tby1 = y1 - th - baseline - 4
                ty   = y1 - baseline - 4
                if tby1 < 0:
                    tby1 = y2 + 2
                    ty   = y2 + th + 2
                cv2.rectangle(draw_frame, (x1, tby1), (x1+tw+2, tby1+th+baseline+2),
                              box_color, cv2.FILLED)
                cv2.putText(draw_frame, display_text, (x1+1, ty),
                            cv2.FONT_HERSHEY_SIMPLEX, font_scale, text_clr, thickness, cv2.LINE_AA)

            rgb = cv2.cvtColor(draw_frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb.shape
            q_img   = QImage(rgb.data, w, h, ch * w, QImage.Format.Format_RGB888)
            pixmap  = QPixmap.fromImage(q_img)
            self.video_label.setPixmap(
                pixmap.scaled(self.video_label.size(),
                              Qt.AspectRatioMode.KeepAspectRatio,
                              Qt.TransformationMode.SmoothTransformation)
            )
            self._manage_emotion_detail_windows(tracker_details)
        except Exception as e:
            print(f"ERROR in frame processing: {type(e).__name__} - {e}")
            traceback.print_exc(limit=2)

        self.update_attendance_table_display()

    def closeEvent(self, event):
        if self.closing_app:
            event.accept()
            return
        print("=" * 30 + " Close Requested " + "=" * 30)
        self.closing_app = True
        if self.timer.isActive():
            self.timer.stop()
        self._close_all_emotion_detail_windows()
        if self.camera_worker and self.camera_worker.isRunning():
            self.camera_worker.stop()
        if self.vid_capture:
            self.vid_capture.release()
            self.vid_capture = None

        user_choice = QMessageBox.StandardButton.No
        if self.recognition_system and OPENPYXL_AVAILABLE:
            user_choice = QMessageBox.question(
                self, "Save Report?",
                "Save final attendance report before closing?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel
            )
            if user_choice == QMessageBox.StandardButton.Cancel:
                self.closing_app = False
                if (not self.timer.isActive() and
                        self.active_camera_source_id is not None):
                    self.timer.start(self.ui_update_delay_ms)
                event.ignore()
                return
            if user_choice == QMessageBox.StandardButton.Yes:
                self.save_report()

        if self.recognition_system:
            self.recognition_system.close()
        event.accept()


# ==============================================================================
# Entry Point
# ==============================================================================
if __name__ == "__main__":
    print("=" * 80)
    print(f"Starting {WINDOW_TITLE_QT}")
    print(f"Python {sys.version.split()[0]}  |  OpenCV {cv2.__version__}")
    try:
        print(f"dlib {dlib.__version__}")
    except AttributeError:
        pass
    print(f"DeepFace available: {DEEPFACE_AVAILABLE}  |  openpyxl: {OPENPYXL_AVAILABLE}")
    print("=" * 80)

    app = QApplication.instance() or QApplication(sys.argv)
    main_window = None
    exit_code = 0
    try:
        main_window = App(KNOWN_FACES_DIR, video_source=LAPTOP_CAM_SRC)
        main_window.show()
        exit_code = app.exec()
    except RuntimeError as e:
        print(f"FATAL: {e}")
        traceback.print_exc()
        exit_code = 1
    except SystemExit as e:
        exit_code = e.code if e.code is not None else 0
    except KeyboardInterrupt:
        print("KeyboardInterrupt — shutting down.")
        if main_window and not main_window.closing_app:
            QTimer.singleShot(0, main_window.close)
            app.processEvents()
        exit_code = 130
    except Exception as e:
        print(f"UNHANDLED EXCEPTION: {type(e).__name__}: {e}")
        traceback.print_exc()
        try:
            QMessageBox.critical(None, "Fatal Error",
                                 f"{type(e).__name__}: {e}\n\nCheck console for details.")
        except Exception:
            pass
        exit_code = 1
    finally:
        if main_window and hasattr(main_window, 'recognition_system') and main_window.recognition_system:
            rs = main_window.recognition_system
            if (hasattr(rs, 'identification_thread') and
                    rs.identification_thread.is_alive()):
                rs.stop_identifier_thread()
                rs.identification_thread.join(1.0)
        if (main_window and main_window.camera_worker and
                main_window.camera_worker.isRunning()):
            main_window.camera_worker.stop()
        print(f"Application exited. {datetime.now()}")
        sys.exit(exit_code)
